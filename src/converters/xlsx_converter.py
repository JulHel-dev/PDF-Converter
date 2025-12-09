"""
Excel/Spreadsheet Converter

Handles Excel and spreadsheet conversions.
Supported: XLSX/XLS/ODS ↔ CSV, JSON, TXT, PDF
"""
import os
from typing import Optional, Dict, List
from src.converters.base_converter import BaseConverter


class XlsxConverter(BaseConverter):
    """
    Handles Excel/spreadsheet conversions.
    
    Supported: XLSX/XLS/ODS ↔ CSV, JSON, TXT, PDF
    """
    
    def __init__(self):
        super().__init__()
        self.supported_input_formats = ['xlsx', 'xls', 'ods']
        self.supported_output_formats = ['csv', 'json', 'txt', 'pdf']
    
    def convert(self, input_path: str, output_path: str, output_format: str) -> bool:
        """Convert spreadsheet to target format."""
        timer_id = self.monitor.start_timer('xlsx_conversion')
        
        try:
            # Step 1: Validate
            self.monitor.log_event('conversion_start', {
                'input': input_path,
                'output_format': output_format,
                'output_path': output_path
            }, severity='INFO', step='Step 1/4: Validating input')
            
            if not self.validate_input(input_path):
                return False
            
            if not self.validate_output(output_path):
                return False
            
            # Step 2: Load spreadsheet
            self.monitor.log_event('spreadsheet_load', {
                'input': input_path
            }, severity='INFO', step='Step 2/4: Loading spreadsheet')
            
            content = self._extract_content(input_path)
            
            if not content:
                return False
            
            # Step 3: Convert to target format
            self.monitor.log_event('format_conversion', {
                'output_format': output_format
            }, severity='INFO', step='Step 3/4: Converting to target format')
            
            result = self._write_output(content, output_path, output_format)
            
            # Step 4: Complete
            self.monitor.log_event('conversion_complete', {
                'input': input_path,
                'output': output_path,
                'success': result
            }, severity='INFO', step='Step 4/4: Conversion complete')
            
            return result
            
        except Exception as e:
            self.monitor.log_event('conversion_failed', {
                'input': input_path,
                'output_format': output_format,
                'error': str(e),
                'error_type': type(e).__name__
            }, severity='ERROR')
            return False
        finally:
            self.monitor.end_timer(timer_id)
    
    def _extract_content(self, xlsx_path: str) -> Optional[Dict]:
        """Extract content from Excel file."""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
            
            content = {
                'sheets': [],
                'properties': {
                    'title': workbook.properties.title,
                    'creator': workbook.properties.creator,
                    'modified': str(workbook.properties.modified) if workbook.properties.modified else None
                }
            }
            
            # Extract each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                sheet_data = {
                    'name': sheet_name,
                    'rows': [],
                    'max_row': sheet.max_row,
                    'max_col': sheet.max_column
                }
                
                # Extract all rows
                for row in sheet.iter_rows(values_only=True):
                    # Convert None to empty string and all values to strings
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    sheet_data['rows'].append(row_data)
                
                content['sheets'].append(sheet_data)
            
            workbook.close()
            return content
            
        except ImportError:
            self.monitor.log_event('dependency_missing', {
                'library': 'openpyxl',
                'install_command': 'pip install openpyxl'
            }, severity='ERROR')
            return None
        except Exception as e:
            self.monitor.log_event('content_extraction_failed', {
                'file': xlsx_path,
                'error': str(e)
            }, severity='ERROR')
            return None
    
    def _write_output(self, content: Dict, output_path: str, format: str) -> bool:
        """Write content to target format."""
        format = format.lower()
        
        try:
            if format == 'csv':
                return self._write_csv(content, output_path)
            elif format == 'json':
                return self._write_json(content, output_path)
            elif format == 'txt':
                return self._write_text(content, output_path)
            elif format == 'pdf':
                return self._write_pdf(content, output_path)
            else:
                self.monitor.log_event('unsupported_output_format', {
                    'format': format
                }, severity='ERROR')
                return False
                
        except Exception as e:
            self.monitor.log_event('output_write_failed', {
                'format': format,
                'error': str(e)
            }, severity='ERROR')
            return False
    
    def _write_csv(self, content: Dict, output_path: str) -> bool:
        """Write content as CSV (first sheet only)."""
        import csv
        
        if not content['sheets']:
            return False
        
        # If multiple sheets, use first sheet and create additional files
        first_sheet = content['sheets'][0]
        
        with open(output_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            for row in first_sheet['rows']:
                writer.writerow(row)
        
        # For multiple sheets, create separate CSV files
        if len(content['sheets']) > 1:
            base_name = os.path.splitext(output_path)[0]
            for i, sheet in enumerate(content['sheets'][1:], start=2):
                sheet_output = f"{base_name}_sheet{i}.csv"
                with open(sheet_output, 'w', encoding='utf-8', newline='') as f:
                    writer = csv.writer(f)
                    for row in sheet['rows']:
                        writer.writerow(row)
        
        return True
    
    def _write_json(self, content: Dict, output_path: str) -> bool:
        """Write content as JSON."""
        import json
        
        # Create JSON structure with all sheets
        json_data = {
            'properties': content['properties'],
            'sheets': []
        }
        
        for sheet in content['sheets']:
            # Convert rows to list of dictionaries using first row as headers
            if sheet['rows']:
                headers = sheet['rows'][0]
                data_rows = []
                
                for row in sheet['rows'][1:]:
                    row_dict = {}
                    for i, value in enumerate(row):
                        col_name = headers[i] if i < len(headers) else f'Column{i+1}'
                        row_dict[col_name] = value
                    data_rows.append(row_dict)
                
                json_data['sheets'].append({
                    'name': sheet['name'],
                    'data': data_rows
                })
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        return True
    
    def _write_text(self, content: Dict, output_path: str) -> bool:
        """Write content as plain text."""
        with open(output_path, 'w', encoding='utf-8') as f:
            # Write properties
            props = content['properties']
            if props.get('title'):
                f.write(f"Title: {props['title']}\n\n")
            
            # Write each sheet
            for sheet in content['sheets']:
                f.write(f"=== Sheet: {sheet['name']} ===\n\n")
                
                for row in sheet['rows']:
                    f.write(' | '.join(row) + '\n')
                
                f.write('\n')
        
        return True
    
    def _write_pdf(self, content: Dict, output_path: str) -> bool:
        """Write content as PDF (simplified version)."""
        try:
            # This is a placeholder - full PDF generation would require
            # libraries like reportlab or conversion through other tools
            self.monitor.log_event('pdf_conversion_note', {
                'message': 'PDF conversion from Excel requires additional setup',
                'suggestion': 'Use Excel, LibreOffice, or specialized conversion tools'
            }, severity='WARNING')
            
            return False  # Not fully implemented
            
        except Exception as e:
            self.monitor.log_event('pdf_conversion_failed', {
                'error': str(e)
            }, severity='ERROR')
            return False
